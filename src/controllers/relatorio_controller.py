import io
import base64
from datetime import date

from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy.orm import joinedload

from models.models import Usuario, Paciente, Sintoma, Avaliacao, SintomaAvaliacao
from controllers.relatorio_stats import (
    montar_query, calcular_kpis, dados_por_mes, dados_por_recomendacao,
    dados_por_sexo, histograma_scores, frequencia_sintomas, por_profissional,
    montar_tabela, FAIXAS_ETARIAS,
)
from controllers.relatorio_charts import gerar_todos

relatorio_bp = Blueprint('relatorio', __name__, url_prefix='/relatorios')


# ---------- helpers ----------

def _filtros_form():
    """Dados auxiliares para os selects/checkboxes da tela de filtros."""
    usuarios = Usuario.query.order_by(Usuario.nome).all() if current_user.is_admin else []
    if current_user.is_admin:
        pacientes = Paciente.query.order_by(Paciente.nome).all()
    else:
        pacientes = Paciente.query.filter_by(id_usuario=current_user.id).order_by(Paciente.nome).all()
    sintomas = Sintoma.query.order_by(Sintoma.label).all()
    return {'usuarios': usuarios, 'pacientes': pacientes, 'sintomas': sintomas,
            'faixas': [f[0] for f in FAIXAS_ETARIAS]}


# ---------- rotas ----------

@relatorio_bp.route('/')
@login_required
def index():
    avaliacoes = montar_query(request.args, current_user).all()
    kpis = calcular_kpis(avaliacoes)
    return render_template('relatorios/index.html',
                           avaliacoes=avaliacoes,
                           kpis=kpis,
                           **_filtros_form())


@relatorio_bp.route('/api/dados')
@login_required
def api_dados():
    avaliacoes = montar_query(request.args, current_user).all()
    return jsonify({
        'kpis': calcular_kpis(avaliacoes),
        'por_mes': dados_por_mes(avaliacoes),
        'recomendacao': dados_por_recomendacao(avaliacoes),
        'sexo': dados_por_sexo(avaliacoes),
        'histograma': histograma_scores(avaliacoes),
        'sintomas': frequencia_sintomas(avaliacoes, top_n=12),
        'profissional': por_profissional(avaliacoes) if current_user.is_admin else [],
    })


@relatorio_bp.route('/export/pdf')
@login_required
def export_pdf():
    anonimizar = request.args.get('anonimizar') in ('1', 'on', 'true')
    avaliacoes = montar_query(request.args, current_user).all()
    kpis = calcular_kpis(avaliacoes)
    tabela = montar_tabela(avaliacoes, anonimizar=anonimizar)
    charts = gerar_todos(avaliacoes, incluir_profissional=current_user.is_admin)
    charts_b64 = {k: base64.b64encode(v).decode('ascii') for k, v in charts.items()}

    html_str = render_template('relatorios/pdf.html',
                               kpis=kpis,
                               tabela=tabela,
                               charts=charts_b64,
                               filtros=request.args,
                               gerado_em=date.today(),
                               gerado_por=current_user.nome,
                               is_admin=current_user.is_admin,
                               anonimizado=anonimizar)
    pdf_bytes = HTML(string=html_str, base_url=request.host_url).write_pdf()
    nome = f"relatorio_triagem_{date.today().isoformat()}.pdf"
    return send_file(io.BytesIO(pdf_bytes), download_name=nome,
                     mimetype='application/pdf', as_attachment=True)


@relatorio_bp.route('/export/xlsx')
@login_required
def export_xlsx():
    anonimizar = request.args.get('anonimizar') in ('1', 'on', 'true')
    # Para a aba "Sintomas" precisamos iterar a.sintomas.sintoma.label sem N+1
    avaliacoes = (montar_query(request.args, current_user)
                  .options(joinedload(Avaliacao.sintomas).joinedload(SintomaAvaliacao.sintoma))
                  .all())
    kpis = calcular_kpis(avaliacoes)
    tabela = montar_tabela(avaliacoes, anonimizar=anonimizar)
    charts = gerar_todos(avaliacoes, incluir_profissional=current_user.is_admin)

    wb = Workbook()

    # ---- Aba Resumo ----
    ws = wb.active
    ws.title = 'Resumo'
    ws['A1'] = 'Relatório de Triagem — Síndrome do X Frágil'
    ws['A1'].font = Font(size=14, bold=True, color='2563EB')
    ws.merge_cells('A1:E1')

    ws['A3'] = 'Gerado em:'
    ws['B3'] = date.today().strftime('%d/%m/%Y')
    ws['A4'] = 'Gerado por:'
    ws['B4'] = current_user.nome

    ws['A6'] = 'Filtros aplicados:'
    ws['A6'].font = Font(bold=True)
    linha = 7
    for k, v in request.args.items():
        if v:
            ws.cell(row=linha, column=1, value=k)
            ws.cell(row=linha, column=2, value=v)
            linha += 1

    linha = max(linha, 8) + 1
    ws.cell(row=linha, column=1, value='KPIs').font = Font(bold=True)
    linha += 1
    for label, valor in [
        ('Total de avaliações', kpis['total']),
        ('Encaminhar', kpis['encaminhar']),
        ('Não encaminhar', kpis['nao_encaminhar']),
        ('% Encaminhamento', f"{kpis['pct_encaminhar']}%"),
        ('Score médio', kpis['score_medio']),
        ('Score mediano', kpis['score_mediano']),
    ]:
        ws.cell(row=linha, column=1, value=label)
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    # Embute graficos na aba Resumo
    linha += 2
    col_chart = 'A'
    for nome_chart, png in charts.items():
        img = XLImage(io.BytesIO(png))
        img.width = 480
        img.height = 220
        ws.add_image(img, f'{col_chart}{linha}')
        linha += 13

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 28

    # ---- Aba Avaliações ----
    ws2 = wb.create_sheet('Avaliações')
    headers = ['ID', 'Data', 'Paciente', 'CPF', 'Sexo', 'Idade', 'Faixa', 'Score', 'Recomendação', 'Profissional']
    bold = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')
    enc_fill = PatternFill('solid', fgColor='FEE2E2')
    nao_fill = PatternFill('solid', fgColor='DCFCE7')

    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')

    for i, row in enumerate(tabela, start=2):
        valores = [
            row['id'], row['data'].strftime('%d/%m/%Y'), row['paciente'], row['cpf'],
            row['sexo'], row['idade'], row['faixa'],
            row['score'], row['recomendacao'], row['profissional'],
        ]
        for col, v in enumerate(valores, start=1):
            c = ws2.cell(row=i, column=col, value=v)
            if col == 9:  # Recomendação
                c.fill = enc_fill if v == 'ENCAMINHAR' else nao_fill

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = ws2.dimensions

    # ---- Aba Sintomas (pivot) ----
    ws3 = wb.create_sheet('Sintomas')
    sintomas = Sintoma.query.order_by(Sintoma.label).all()
    for j, header in enumerate(['Avaliação', 'Data', 'Paciente'], start=1):
        c = ws3.cell(row=1, column=j, value=header)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    for j, s in enumerate(sintomas, start=4):
        c = ws3.cell(row=1, column=j, value=s.label)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    from controllers.relatorio_stats import _mascarar_nome
    for i, a in enumerate(avaliacoes, start=2):
        ws3.cell(row=i, column=1, value=a.id)
        ws3.cell(row=i, column=2, value=a.data.strftime('%d/%m/%Y'))
        nome_p = _mascarar_nome(a.paciente.nome) if anonimizar else a.paciente.nome
        ws3.cell(row=i, column=3, value=nome_p)
        marcados = {sa.id_sintoma for sa in a.sintomas if sa.presente}
        for j, s in enumerate(sintomas, start=4):
            ws3.cell(row=i, column=j, value='✔' if s.id in marcados else '')
    ws3.freeze_panes = 'D2'

    # Saida
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"relatorio_triagem_{date.today().isoformat()}.xlsx"
    return send_file(buf, download_name=nome, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
